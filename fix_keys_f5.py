#!/usr/bin/env python3
"""Reemplaza las claves incorrectas de FECHA 5 usando los nombres del fixture JSON."""

import openpyxl, json, re, unicodedata, os

EXCEL  = "FECHA 5 (1).xlsx"
DATAJS = "data.js"

SHEET_SECTION = {
    "Inter":  "intermedia",
    "Pre A":  "preinter_a",
    "Pre B":  "preinter_b",
    "Pre C":  "preinter_c",
    "Pre D":  "preinter_d",
    "Pre E":  "preinter_e",
    "Pre F":  "preinter_f",
    "M22":    "m22",
}

SECTION_FIXTURE = {
    "intermedia":  "urba-intermedia.json",
    "preinter_a":  "urba-preinter-a.json",
    "preinter_b":  "urba-preinter-b.json",
    "preinter_c":  "urba-preinter-c.json",
    "preinter_d":  "urba-preinter-d.json",
    "preinter_e":  "urba-preinter-e.json",
    "preinter_f":  "urba-preinter-f.json",
    "m22":         "urba-m22.json",
}

COMPOUND_FIRST_NAMES = {
    ("JUAN","PEDRO"),("JUAN","PABLO"),("JUAN","JOSE"),("JUAN","CRUZ"),
    ("JUAN","MANUEL"),("JUAN","CARLOS"),("JUAN","IGNACIO"),("JUAN","BAUTISTA"),
    ("JUAN","MARTIN"),("JUAN","FRANCISCO"),("JUAN","MARCOS"),("JUAN","SEGUNDO"),
    ("JOSE","MARIA"),("JOSE","IGNACIO"),
}

# ── utilidades ─────────────────────────────────────────────────────────────────

def rm_accents(s):
    return "".join(c for c in unicodedata.normalize("NFD",s)
                   if unicodedata.category(c) != "Mn")

def make_id(s):
    s = rm_accents(s.strip().lower())
    s = re.sub(r"[^a-z0-9]+","_",s)
    return s.strip("_")

def canon(name):
    """Normalización canónica para comparar nombres de equipo."""
    n = rm_accents(name.strip().lower())
    n = re.sub(r"[^a-z0-9\s]","",n).strip()
    # Reemplazar abreviaturas conocidas
    n = n.replace("bacrc","buenos aires c rc")
    n = n.replace("c&rc","c rc")
    # Quitar sufijos de división al final
    n = re.sub(r"\s+(d|e|f|g|h|b|c|m22|m22\s*b|m22\s*c)\s*$","",n).strip()
    return n

def similarity(a, b):
    aw, bw = set(canon(a).split()), set(canon(b).split())
    if not aw or not bw: return 0
    return len(aw & bw) / len(aw | bw)

def best_fixture_match(excel_home, excel_away, fixture_matches):
    best_score, best = -1, None
    for fm in fixture_matches:
        fh, fa = fm["home"], fm["away"]
        score  = similarity(excel_home, fh) + similarity(excel_away, fa)
        score2 = similarity(excel_home, fa) + similarity(excel_away, fh)
        s = max(score, score2)
        if s > best_score:
            best_score, best = s, fm
    return best

# ── normalización de nombres de jugadores ─────────────────────────────────────

def title_tok(s):
    if not s: return s
    LOWER = {"de","del","la","las","los","el","y","van","von","da","di"}
    if s.lower() in LOWER: return s.lower()
    if "-" in s and len(s) > 2: return "-".join(p.capitalize() for p in s.split("-"))
    return s[0].upper() + s[1:].lower() if len(s)>1 else s.upper()

def is_upper_tok(s):
    letters = re.sub(r"[^a-zA-ZáéíóúÁÉÍÓÚüÜñÑ]","",s)
    return bool(letters) and letters == letters.upper()

def normalize_name(raw):
    if raw is None: return None
    raw = str(raw).strip()
    if raw.startswith("="): return None
    if "SIN INFORMACION" in raw.upper(): return "SIN INFORMACION"
    raw = re.sub(r"\s*\(C\)\s*$","",raw,flags=re.IGNORECASE).strip()
    raw = re.sub(r"\s*\(C$","",raw).strip()
    raw = re.sub(r"^\d+\s*[-–.]\s*","",raw).strip()
    if not raw: return None
    if "," in raw:
        parts = [p.strip() for p in raw.split(",",1)]
        ap  = " ".join(title_tok(w) for w in parts[0].split())
        nom = " ".join(title_tok(w) for w in parts[1].split()) if len(parts)>1 else ""
        return f"{ap}, {nom}" if nom else ap
    tokens = raw.split()
    if not tokens: return raw
    if len(tokens) == 1: return title_tok(tokens[0])
    def is_ini(t): return bool(re.match(r"^[A-ZÁÉÍÓÚ]\.$",t,re.IGNORECASE))
    if is_ini(tokens[0]):
        ini, rs = tokens[0].upper(), 1
        while rs < len(tokens) and is_ini(tokens[rs]):
            ini += " " + tokens[rs].upper(); rs += 1
        ap = " ".join(title_tok(t) for t in tokens[rs:])
        return f"{ap}, {ini}" if ap else ini
    up = [is_upper_tok(t) for t in tokens]
    first_up, last_up = up[0], up[-1]
    if first_up and not last_up:
        split = sum(1 for u in up if u) if all(up[:up.index(False)]) else next(i for i,u in enumerate(up) if not u)
        ap  = " ".join(title_tok(t) for t in tokens[:split])
        nom = " ".join(title_tok(t) for t in tokens[split:])
        return f"{ap}, {nom}"
    if not first_up and last_up:
        split = len(tokens)
        for i in range(len(tokens)-1,-1,-1):
            if up[i]: split = i
            else: break
        nom = " ".join(title_tok(t) for t in tokens[:split])
        ap  = " ".join(title_tok(t) for t in tokens[split:])
        return f"{ap}, {nom}"
    if all(up):
        n = len(tokens)
        if n == 2: return f"{title_tok(tokens[0])}, {title_tok(tokens[1])}"
        last2 = (tokens[-2], tokens[-1])
        if last2 in COMPOUND_FIRST_NAMES:
            ap  = " ".join(title_tok(t) for t in tokens[:-2])
            nom = " ".join(title_tok(t) for t in tokens[-2:])
            return f"{ap}, {nom}"
        if len(tokens[-1])==1 and tokens[-1].isalpha():
            ap  = title_tok(tokens[0])
            nom = " ".join(title_tok(t) for t in tokens[1:-1]) + " " + tokens[-1].upper() + "."
            return f"{ap}, {nom.strip()}"
        # Detect leading single-letter initials
        li = 0
        for t in tokens:
            if len(re.sub(r"[^a-zA-Z]","",t)) == 1: li += 1
            else: break
        if 0 < li < n:
            ini = " ".join(t.upper().rstrip(".")+"." for t in tokens[:li])
            ap  = " ".join(title_tok(t) for t in tokens[li:])
            return f"{ap}, {ini}"
        ap  = " ".join(title_tok(t) for t in tokens[:-1])
        nom = title_tok(tokens[-1])
        return f"{ap}, {nom}"
    if len(tokens)==2: return f"{title_tok(tokens[1])}, {title_tok(tokens[0])}"
    nom = " ".join(title_tok(t) for t in tokens[:-1])
    ap  = title_tok(tokens[-1])
    return f"{ap}, {nom}"

# ── lectura Excel ──────────────────────────────────────────────────────────────

def parse_sheet(ws):
    rows   = list(ws.iter_rows(min_row=1, max_row=17, values_only=True))
    hdr    = rows[0]
    teams  = rows[1]
    players= rows[2:17]
    partidos = []
    col = 0
    while col < len(hdr):
        if hdr[col] and str(hdr[col]).strip().upper().startswith("LOCAL"):
            lc, vc = col, col+1
            ln = teams[lc] if lc < len(teams) else None
            vn = teams[vc] if vc < len(teams) else None
            if ln and vn:
                ln = str(ln).strip(); vn = str(vn).strip()
                lj = [normalize_name(pr[lc] if lc<len(pr) else None) for pr in players]
                vj = [normalize_name(pr[vc] if vc<len(pr) else None) for pr in players]
                partidos.append({"xl_local":ln,"xl_visit":vn,"lj":lj,"vj":vj})
            col += 3
        else:
            col += 1
    return partidos

# ── generación JS ──────────────────────────────────────────────────────────────

def jugadores_js(jugs, indent="          "):
    nombres = [j for j in jugs if j is not None]
    if any("SIN INFORMACION" in (j or "") for j in nombres):
        return f'{indent}"Formación no disponible"'
    lineas = []
    for i, j in enumerate([j for j in nombres if j and "SIN INFORMACION" not in j],1):
        lineas.append(f'{indent}"{i}. {j}"')
    return ",\n".join(lineas)

def gen_entry(match_key, local_name, visit_name, lj, vj):
    ljs = jugadores_js(lj)
    vjs = jugadores_js(vj)
    lines = [
        f"    {match_key}: {{",
        f"      local: {{",
        f'        nombre: "{local_name}",',
        f"        jugadores: [",
        ljs + ",",
        f"        ],",
        f"        cambios: [],",
        f"      }},",
        f"      visitante: {{",
        f'        nombre: "{visit_name}",',
        f"        jugadores: [",
        vjs + ",",
        f"        ],",
        f"        cambios: [],",
        f"      }},",
        f"    }},",
        "",
    ]
    return "\n".join(lines)

# ── reemplazar en data.js ──────────────────────────────────────────────────────

def remove_fecha5_block(content, section_key):
    """Elimina todo el bloque FECHA 5 de una sección."""
    sec_start = content.find(f"\n  {section_key}: {{")
    if sec_start == -1: return content, False
    f5_pos = content.find("    // ── FECHA 5", sec_start)
    if f5_pos == -1: return content, False
    # Encontrar cierre de la sección (2-space indent },)
    # Contar profundidad desde sec_start
    sec_open = content.index("{", sec_start + len(f"\n  {section_key}: "))
    depth = 1; i = sec_open + 1
    while i < len(content) and depth > 0:
        if content[i] == "{": depth += 1
        elif content[i] == "}": depth -= 1
        i += 1
    sec_close = i - 1  # posición del } de cierre
    # Borrar desde f5_pos hasta sec_close (excluyendo el } de cierre)
    new_content = content[:f5_pos] + content[sec_close:]
    return new_content, True

def insert_before_close(content, section_key, new_text):
    sec_start = content.find(f"\n  {section_key}: {{")
    sec_open  = content.index("{", sec_start + len(f"\n  {section_key}: "))
    depth = 1; i = sec_open + 1
    while i < len(content) and depth > 0:
        if content[i] == "{": depth += 1
        elif content[i] == "}": depth -= 1
        i += 1
    sec_close = i - 1
    return content[:sec_close] + "\n" + new_text + content[sec_close:]

# ── main ───────────────────────────────────────────────────────────────────────

def main():
    wb = openpyxl.load_workbook(EXCEL)
    with open(DATAJS, encoding="utf-8") as f:
        content = f.read()

    for sheet_name, section_key in SHEET_SECTION.items():
        if sheet_name not in wb.sheetnames:
            print(f"SKIP {sheet_name}: no existe en Excel")
            continue

        fixture_file = "data/" + SECTION_FIXTURE[section_key]
        if not os.path.exists(fixture_file):
            print(f"SKIP {section_key}: no existe {fixture_file}")
            continue

        # Leer fixture round 5
        fdata   = json.load(open(fixture_file))
        round5  = fdata["matches"].get("5")
        if not round5:
            print(f"SKIP {section_key}: no hay round 5 en fixture")
            continue
        fix_matches = round5["ms"]

        # Leer Excel
        ws       = wb[sheet_name]
        partidos = parse_sheet(ws)

        # Eliminar bloque FECHA 5 existente
        content, removed = remove_fecha5_block(content, section_key)
        if removed:
            print(f"  Eliminado bloque FECHA 5 de {section_key}")

        # Generar nuevas entradas usando claves del fixture
        entries_text = "    // ── FECHA 5 ────────────────────────────────────────────────\n"

        for p in partidos:
            fm = best_fixture_match(p["xl_local"], p["xl_visit"], fix_matches)
            if fm is None:
                print(f"  WARNING: no match found for {p['xl_local']} vs {p['xl_visit']}")
                continue

            # Clave derivada de nombres del fixture
            match_key   = make_id(fm["home"]) + "_" + make_id(fm["away"])
            local_name  = fm["home"]
            visit_name  = fm["away"]

            # Si el local/visitante está invertido respecto al Excel, intercambiar jugadores
            if similarity(p["xl_local"], fm["home"]) < similarity(p["xl_local"], fm["away"]):
                lj, vj = p["vj"], p["lj"]
                local_name, visit_name = fm["away"], fm["home"]
                match_key = make_id(fm["away"]) + "_" + make_id(fm["home"])
            else:
                lj, vj = p["lj"], p["vj"]

            entries_text += gen_entry(match_key, local_name, visit_name, lj, vj)
            print(f"  ✓ {section_key}: {match_key}")

        content = insert_before_close(content, section_key, entries_text)

    with open(DATAJS, "w", encoding="utf-8") as f:
        f.write(content)
    print("\n¡data.js actualizado!")

if __name__ == "__main__":
    main()
